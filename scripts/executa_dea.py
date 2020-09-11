"""
Executa a análise DEA separadamente em cada cluster e armazena o resultado em planilhas
individualizadas por cluster e em uma única planilha consolidada.
"""

import os

import pandas as pd
import numpy as np
import xlsxwriter
import openpyxl

from pydea_wrapper import ModeloDEA
import consts



# Parâmetros do modelo do TCU
input_categories = consts.COLUNAS_DEA[:4]
output_categories = consts.COLUNAS_DEA[4:]
virtual_weight_restrictions = ['CNES_LEITOS_SUS >= 0.16',
                               'CNES_SALAS >= 0.09',
                               'CNES_MEDICOS >= 0.16',
                               'CNES_PROFISSIONAIS_ENFERMAGEM >= 0.09']


def _cria_arquivo_xlsx(first_period, last_period):

    # Tenta abrir arquivo xlsx "resultados_dea_'first_period'_'last_period'" caso exista
    try:

        workbook = \
            openpyxl.load_workbook(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultados_dea_{}_{}.xlsx'.format(first_period, last_period)))

        workbook.save(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultados_dea_{}_{}.xlsx'.format(first_period, last_period)))

    # Cria arquivo xlsx "resultados_dea_'first_period'_'last_period'" caso não exista
    except:

        workbook = \
            xlsxwriter.Workbook(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultados_dea_{}_{}.xlsx'.format(first_period, last_period)))

        workbook.close()


def _salva_em_planilha_xlsx(first_period, last_period, year, month):

    # Abre o arquivo xlsx "resultados"
    workbook = \
        openpyxl.load_workbook(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultados_dea_{}_{}.xlsx'.format(first_period, last_period)))

    # Cria um objeto ExcelWriter para escrever no arquivo xlsx "resultados_dea_'first_period'_'last_period'"
    writer = \
        pd.ExcelWriter(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultados_dea_{}_{}.xlsx'.format(first_period, last_period)),
                       engine='openpyxl')

    # Vincula o objeto "workbook" ao atributo "book" de "writer"
    writer.book = workbook

    # Lê o arquivo xlsx "resultado_dados_tratados_dea_'ano'_'mes'" como um objeto pandas DataFrame
    df = pd.read_excel(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultado_dados_tratados_dea_{}_{}.xlsx'.format(year, month)))

    # Colocaos dados de parte das colunas de "df" na planilha 'ano'-'mes' do arquivo xlsx "resultados_dea_'first_period'_'last_period'"
    df[consts.COLUNAS_CATEGORICAS + ['EFICIENCIA' + 'EFICIENCIA_NORMALIZADA']].to_excel(writer, sheet_name='{}-{}'.format(year, month), index=False)

    # Salva os dados
    writer.save()

    # Fecha o arquivo xlsx
    writer.close()


def executa_dea(year=2019, month=False):

    # Instancia modelo DEA
    modelo_tcu = ModeloDEA(input_categories, output_categories, virtual_weight_restrictions)

    if month:

        # Coleta path do arquivo "xlsx" de dados com a coluna de label do cluster
        arquivo_dados_dea = os.path.join(consts.DIRETORIO_DADOS, 'dados_tratados_dea_{}_{}.xlsx'.format(year, month))

    else:

        # Coleta path do arquivo "xlsx" de dados com a coluna de label do cluster
        arquivo_dados_dea = os.path.join(consts.DIRETORIO_DADOS, 'dados_tratados_dea_{}.xlsx'.format(year, month))

    # Executa DEA
    modelo_tcu.executa_por_cluster(arquivo_dados_dea, consts.DIRETORIO_RESULTADOS, 'CLUSTER')

    # Deleta o arquivo xlsx de dados tratados para execução da DEA
    os.unlink(arquivo_dados_dea)


def executa_deas(first_period='01-2017', last_period='12-2019'):

    _cria_arquivo_xlsx(first_period, last_period)

    # Coleta a substring referente ao ano e a transforma no tipo int
    first_year = int(first_period[3:])

    # Coleta a substring referente ao ano e a transforma no tipo int
    last_year = int(last_period[3:])

    # Coleta a substring do mês
    last_month = last_period[:2]

    # Itera sobre os anos
    for ano in np.arange(first_year, last_year + 1):

        if ano == last_year:

            # Itera sobre os meses do último ano de dados
            for mes in np.arange(1, int(last_month) + 1):

                # Converte o tipo int referenciado à variável "mes" como tipo string
                mes = str(mes)

                # Preenche com zeros à esquerda a string "mes" até ficar com dois dígitos
                mes = mes.zfill(2)

                executa_dea(ano, mes)

                _salva_em_planilha_xlsx(first_period, last_period, ano, mes)

        else:

            # Itera sobre os meses de todos os anos menos do último
            for mes in np.arange(1, 13):

                mes = str(mes)
                mes = mes.zfill(2)

                executa_dea(ano, mes)

                _salva_em_planilha_xlsx(first_period, last_period, ano, mes)

    # Abre o arquivo xlsx "resultado_dados_tratados_dea_'ano'_'mes'"
    workbook = \
        openpyxl.load_workbook(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultado_dados_tratados_dea_{}_{}.xlsx'.format(year, month)))

    # Acessa o atributo "sheetnames" de "workbook"
    sheets_names = workbook.sheetnames

    # Deleta a planilha de nome "Sheet1"
    workbook.remove(workbook.get_sheet_by_name('Sheet1'))

    # Salva os dados
    workbook.save(os.path.join(consts.DIRETORIO_RESULTADOS, 'resultado_dados_tratados_dea_{}_{}.xlsx'.format(year, month)))



if __name__ == '__main__':

    executa_dea(2019)
    #executa_deas('01-2017', '12-2019')
